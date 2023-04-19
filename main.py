import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.feature_selection import SelectKBest, f_classif
plt.style.use('fivethirtyeight')
plt.rcParams['figure.figsize'] = (11, 7)


ant_only_bugs_file = 'datasets/ant-ivy-all-versions-only-bugs.xlsx'
start_row_only_bugs = 4

ant_all_versions_file = 'datasets/ant-ivy-all-versions.xlsx'
start_row_all_versions = 8

c1 = ['0-188', '0-82', '0-70']
c2 = ['3-0', '2-2']

c331 = ['1-21', '1-25', '2-27']
c332 = ['3-21']
c333 = ['2-11']
c334 = ['2-23']
c335 = ['1-9']

c31 = ['2-34']
c32 = ['5-32', '5-17', '4-44']
c33 = c331 + c332 + c333 + c334 + c335

c3 = c31 + c32 + c33
c4 = []


def analyzeDefectsVsNonDefects():

    worksheet = openpyxl.load_workbook(ant_all_versions_file, data_only=True).active

    df_all_versions = pd.DataFrame(worksheet.values)
    df_all_versions.reset_index(inplace=True, drop=True)

    df_all_versions.drop([0, 1, 2, 3, 4, 5, 6, 7, 8], inplace=True)
    df_all_versions.drop(df_all_versions.columns[[0, 1, 2, 3]], axis=1, inplace=True)
    print(df_all_versions.head(5))

    Y = df_all_versions.iloc[:, 4]
    print(Y)

    X = df_all_versions.iloc[:, df_all_versions.columns != 4]
    print(X)

    # apply SelectKBest class to order the features by importance
    best_features = SelectKBest(score_func=f_classif, k='all')
    fit = best_features.fit(X, Y)

    df_scores = pd.DataFrame(fit.scores_)
    df_columns = pd.DataFrame(X.columns)

    # concat two dataframes for better visualization
    featureScores = pd.concat([df_columns, df_scores], axis=1)
    featureScores.columns = ['Specs', 'Score']

    featureScores['Specs'] = featureScores['Specs'] - 4
    featureScores = featureScores.fillna(-1)

    featureScores.to_csv('defects-vs-nonDefects/feature-scores.csv', index=False)

    print('Best 20 scores are: ', featureScores.nlargest(20, 'Score'))

def plotBestFeaturesDefectsVsAll():
    df_results_clean = pd.read_csv('defects-vs-nonDefects/feature-scores-cleaned-up.csv')
    print('Best 20 scores are: ', df_results_clean.nlargest(20, 'Score'))

    # creating the bar plot
    plt.bar(df_results_clean['Specs'], df_results_clean['Score'])
    plt.xlabel("Features")
    plt.ylabel("Feature scores")
    plt.savefig('defects-vs-nonDefects/best-features.png')
    plt.cla()
    plt.clf()

    plt.bar(df_results_clean['Specs'].head(10), df_results_clean['Score'][:10])
    plt.xlabel("Features")
    plt.ylabel("Feature scores")
    plt.savefig('defects-vs-nonDefects/best-10-features.png')


def buildC4(iWorksheet):
    for i in range(start_row_only_bugs + 2, iWorksheet.max_row):
        som_code = iWorksheet.cell(i, 4).value
        if som_code not in c1 or som_code not in c2 or som_code not in c3:
            c4.append(som_code)

def markClasses(iCluster, iDataFrame):
    labels = []

    for index, row in iDataFrame.iterrows():
        if row[3] in iCluster:
            labels.append(1)
        else: labels.append(0)

    iDataFrame['Labels'] = labels

    return iDataFrame

def analyzeClusterVsDefects(iDataFIle, iCluster, iClusterName, iDir):
    worksheet = openpyxl.load_workbook(iDataFIle, data_only=True).active

    buildC4(worksheet)

    data_frame = pd.DataFrame(worksheet.values)
    data_frame.reset_index(inplace=True, drop=True)
    data_frame.drop([0, 1, 2, 3, 4, 5, 6, 7, 8], inplace=True)

    data_frame = markClasses(iCluster, data_frame)

    data_frame.drop(data_frame.columns[[0, 1, 2, 3]], axis=1, inplace=True)

    Y = data_frame['Labels']
    X = data_frame.iloc[:, data_frame.columns != 'Labels']

    # apply SelectKBest class to order the features by importance
    best_features = SelectKBest(score_func=f_classif, k='all')
    fit = best_features.fit(X, Y)

    df_scores = pd.DataFrame(fit.scores_)
    df_columns = pd.DataFrame(X.columns)

    # concat two dataframes for better visualization
    featureScores = pd.concat([df_columns, df_scores], axis=1)
    featureScores.columns = ['Specs', 'Score']

    featureScores['Specs'] = featureScores['Specs'] - 4
    featureScores = featureScores.fillna(-1)


    path = iDir + f'/feature-scores-cluster-{iClusterName}.csv'
    featureScores.to_csv(path, index=False)

    print('Best 20 scores are:\n', featureScores.nlargest(20, 'Score'))

if __name__ == "__main__":
    # analyzeDefectsVsNonDefects()
    # plotBestFeaturesDectsVsAll()

    # analyzeClusterVsDefects(ant_only_bugs_file, c33, 'c33', 'cluster-vs-defects')
    # analyzeClusterVsDefects(ant_only_bugs_file, c3, 'c3', 'cluster-vs-defects')
    # analyzeClusterVsDefects(ant_only_bugs_file, c4, 'c4', 'cluster-vs-defects')

    # analyzeClusterVsDefects(ant_all_versions_file, c33, 'c33',  'cluster-vs-all')
    # analyzeClusterVsDefects(ant_all_versions_file, c3, 'c3', 'cluster-vs-all')
    analyzeClusterVsDefects(ant_all_versions_file, c4, 'c4', 'cluster-vs-all')



